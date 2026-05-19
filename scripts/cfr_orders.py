"""
cfr_orders.py
SAP Order Processor — web version (tkinter removed)
Change vs original: REJECTED rows are fully colored red in the output Excel.
"""

import re
import csv
import io
import pandas as pd
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from zoneinfo import ZoneInfo
    _TZ = ZoneInfo("Europe/Brussels")
except Exception:
    _TZ = None


def now_local() -> datetime:
    return datetime.now(_TZ) if _TZ else datetime.now()


def brussels_stamp() -> str:
    return now_local().strftime("%Y%m%dT%H%M")


def infer_year_from_mmdd(mm: int, dd: int, forced_year: Optional[int] = None) -> int:
    n = now_local()
    if forced_year is not None:
        return forced_year
    y = n.year
    if mm > n.month:
        y -= 1
    return y


def extract_date_from_ref(ref: str, year: Optional[int] = None) -> str:
    if not isinstance(ref, str):
        return ""
    m = re.search(r"-\s*([0-9]{4})\b", ref)
    if not m:
        return ""
    mmdd = m.group(1)
    mm = int(mmdd[0:2])
    dd = int(mmdd[2:4])
    y = infer_year_from_mmdd(mm, dd, forced_year=year)
    try:
        d = datetime(y, mm, dd)
        return d.strftime("%d.%m.%Y")
    except ValueError:
        return ""


def extract_ean(desc: str) -> str:
    if not isinstance(desc, str):
        return ""
    s = str(desc).strip()
    m = re.search(r"\b(5\d{12})\b", s)
    if m:
        return m.group(1)
    m = re.search(r"(5\d{12})", s)
    if m:
        return m.group(1)
    return ""


def to_number(x) -> float:
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        m = re.search(r"[-+]?\d*\.?\d+", s)
        return float(m.group(0)) if m else 0.0


def load_excel_flex(file_obj) -> pd.DataFrame:
    """Load from file-like object."""
    try:
        return pd.read_excel(file_obj, header=None, dtype=str, engine="openpyxl")
    except Exception:
        file_obj.seek(0)
        for enc in ("utf-8-sig", "cp1252", "latin1"):
            try:
                return pd.read_csv(file_obj, header=None, dtype=str,
                                   engine="python", sep=None, encoding=enc)
            except Exception:
                file_obj.seek(0)
                for sep in (";", ",", "\t", "|"):
                    try:
                        return pd.read_csv(file_obj, header=None, dtype=str,
                                           sep=sep, encoding=enc)
                    except Exception:
                        file_obj.seek(0)
        return pd.read_csv(file_obj, header=None, dtype=str)


def transform_for_sap(rows, assumed_year: Optional[int] = None):
    if isinstance(rows, pd.DataFrame):
        seq = rows.fillna("").values.tolist()
    else:
        seq = []
        for raw in rows:
            try:
                seq.append(list(raw))
            except TypeError:
                continue

    idx_ref  = 0
    idx_addr = 4
    idx_post = 5
    idx_city = 6
    idx_desc = 13
    idx_ord  = 14
    idx_recv = 18
    idx_unit = 19
    start_i  = 0

    if len(seq) > 0 and isinstance(seq[0], (list, tuple)):
        hdr = seq[0]
        hdr_norm = [str(x).strip().lower() for x in hdr]

        def find(names):
            for j, s in enumerate(hdr_norm):
                if s in names:
                    return j
            return None

        found = False
        t = find({"customer reference","order reference","orderref","customer ref","customer referentie","reference"})
        if t is not None: idx_ref = t; found = True
        t = find({"receiver address","address","receiver adress","receiver adres"})
        if t is not None: idx_addr = t; found = True
        t = find({"receiver city","city"})
        if t is not None: idx_city = t; found = True
        t = find({"postcode","post code","postal code","zip","zip code","receiver postcode"})
        if t is not None: idx_post = t; found = True
        t = find({"product description","description","desc","product desc","artikelomschrijving"})
        if t is not None: idx_desc = t; found = True
        t = find({"ordered","ordered qty","qty ordered","quantity ordered"})
        if t is not None: idx_ord = t; found = True
        t = find({"received","qty received","quantity received","received qty","ontvangen"})
        if t is not None: idx_recv = t; found = True
        t = find({"unit","uom"})
        if t is not None: idx_unit = t; found = True
        if found:
            start_i = 1

    max_idx = max(idx_ref, idx_addr, idx_post, idx_city, idx_desc, idx_ord, idx_recv, idx_unit) + 1

    def pad(row):
        if len(row) < max_idx:
            row = list(row) + [""] * (max_idx - len(row))
        return row

    out = []
    for i in range(start_i, len(seq)):
        row = seq[i]
        if not isinstance(row, (list, tuple)):
            continue
        row = pad(row)

        order_ref = "" if row[idx_ref]  is None else str(row[idx_ref])
        addr      = "" if row[idx_addr] is None else str(row[idx_addr])
        postcode  = "" if row[idx_post] is None else str(row[idx_post])
        city      = "" if row[idx_city] is None else str(row[idx_city])
        desc      = "" if row[idx_desc] is None else str(row[idx_desc])

        qty_ord  = int(round(to_number(row[idx_ord])))
        qty_recv = int(round(to_number(row[idx_recv])))

        ean = extract_ean(desc)
        if ean == "":
            for cell in row:
                if cell is None:
                    continue
                m = re.search(r"(5\d{12})", str(cell))
                if m:
                    ean = m.group(1)
                    break

        if order_ref.strip() == "" and ean == "" and qty_ord == 0 and qty_recv == 0:
            continue

        order_date = extract_date_from_ref(order_ref, assumed_year)

        if qty_recv > 0:
            out.append({
                "OrderRef":           order_ref.strip(),
                "OrderDate":          order_date,
                "SoldTo(Postcode)":   postcode.strip(),
                "ReceiverCity":       city.strip(),
                "Address":            addr.strip(),
                "EAN":                ean,
                "Qty":                qty_recv,
                "Unit":               "PC",
                "DUMMY":              "DUMMY",
                "LineType":           "ACCEPTED",
                "OriginalQtyOrdered": qty_ord,
                "QtyReceived":        qty_recv,
                "QtyRejected":        max(qty_ord - qty_recv, 0),
            })

        remainder = qty_ord - qty_recv
        if remainder > 0:
            out.append({
                "OrderRef":           order_ref.strip(),
                "OrderDate":          order_date,
                "SoldTo(Postcode)":   postcode.strip(),
                "ReceiverCity":       city.strip(),
                "Address":            addr.strip(),
                "EAN":                ean,
                "Qty":                remainder,
                "Unit":               "PC",
                "DUMMY":              "DUMMY",
                "LineType":           "REJECTED",
                "OriginalQtyOrdered": qty_ord,
                "QtyReceived":        qty_recv,
                "QtyRejected":        remainder,
            })

    cols = ["OrderRef","OrderDate","SoldTo(Postcode)","ReceiverCity","Address",
            "EAN","Qty","Unit","DUMMY","LineType",
            "OriginalQtyOrdered","QtyReceived","QtyRejected"]
    out_df = pd.DataFrame(out, columns=cols)

    if not out_df.empty:
        for c in ["Qty","OriginalQtyOrdered","QtyReceived","QtyRejected"]:
            out_df[c] = pd.to_numeric(out_df[c], errors="coerce").fillna(0).astype(int)
        out_df.sort_values(by=["OrderRef","LineType","EAN"], inplace=True)

    return out_df


def save_as_xlsx(out_df: pd.DataFrame) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import numbers as xl_numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "SAP"

    headers = list(out_df.columns)
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = header_font
    ws.freeze_panes = "A2"

    title_fill    = PatternFill("solid", fgColor="DDDDDD")
    title_font    = Font(bold=True)
    wrap          = Alignment(wrap_text=True, vertical="center")
    thick_top     = Border(top=Side(style="thick"))
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    # ── NEW: red fill for REJECTED rows ──────────────────────
    red_fill = PatternFill("solid", fgColor="FFCCCC")
    red_font = Font(color="C0392B")

    INT_COLS  = {"Qty","OriginalQtyOrdered","QtyReceived","QtyRejected"}
    TEXT_COLS = {"OrderRef","OrderDate","SoldTo(Postcode)","ReceiverCity",
                 "Address","EAN","Unit","DUMMY","LineType"}

    current_row = 1

    for order_ref, g in out_df.groupby("OrderRef", sort=False):
        first = g.iloc[0]
        title_text = (
            f"Order {first['OrderRef']} — Date {first['OrderDate']} — "
            f"SoldTo {first['SoldTo(Postcode)']} — Address {first['Address']}"
        )

        current_row += 1
        ws.insert_rows(current_row)

        ws.cell(row=current_row, column=1, value=title_text)
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row,   end_column=len(headers)
        )
        tc = ws.cell(row=current_row, column=1)
        tc.fill      = title_fill
        tc.font      = title_font
        tc.alignment = wrap
        tc.border    = thick_top

        for _, r in g.iterrows():
            current_row += 1
            is_rejected = str(r.get("LineType","")).strip().upper() == "REJECTED"

            row_vals = []
            for h in headers:
                v = r[h]
                if h in INT_COLS:
                    try:    v = int(round(float(v)))
                    except: v = 0
                else:
                    v = "" if pd.isna(v) else str(v)
                row_vals.append(v)

            ws.append(row_vals)

            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                # Number format
                if h in INT_COLS:
                    cell.number_format = "0"
                elif h in TEXT_COLS:
                    cell.number_format = xl_numbers.FORMAT_TEXT
                else:
                    cell.number_format = "General"
                cell.border = thin_border
                # ── Red color for REJECTED ─────────────────
                if is_rejected:
                    cell.fill = red_fill
                    cell.font = red_font

    widths = {"A":18,"B":14,"C":16,"D":20,"E":45,"F":18,"G":10,"H":8,"I":10,"J":12,"K":20,"L":14,"M":14}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run_cfr_orders(file_obj, assumed_year: Optional[int] = None):
    """
    Main entry point for the web module.
    Returns (buf: BytesIO, stats: dict)
    """
    df     = load_excel_flex(file_obj)
    out_df = transform_for_sap(df, assumed_year=assumed_year)
    buf    = save_as_xlsx(out_df)

    accepted = int((out_df["LineType"] == "ACCEPTED").sum()) if not out_df.empty else 0
    rejected = int((out_df["LineType"] == "REJECTED").sum()) if not out_df.empty else 0
    orders   = int(out_df["OrderRef"].nunique())             if not out_df.empty else 0

    stats = {
        "orders":   orders,
        "accepted": accepted,
        "rejected": rejected,
        "total":    len(out_df),
    }
    return buf, stats
