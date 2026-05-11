"""
promo_uplift_calc.py — optimized version with fast reading
"""

import pandas as pd
import numpy as np
import io
import re
from datetime import datetime, timedelta


KNOWN_DIM_COLS = {
    'chain', 'pet', 'subtype', 'productdescription', 'product description',
    'productid', 'product id', 'country', 'ean', 'material', 'sku', 'description'
}

def get_dim_cols(df):
    return [c for c in df.columns if str(c).strip().lower() in KNOWN_DIM_COLS]

def get_week_cols(df):
    dims = set(get_dim_cols(df))
    return [c for c in df.columns
            if c not in dims and re.search(r'\d{2}/\d{2}/\d{4}', str(c).strip())]

def parse_date(col):
    try:
        return datetime.strptime(str(col).strip(), "%d/%m/%Y").date()
    except Exception:
        return None

def find_col(df, candidates):
    col_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in col_map:
            return col_map[cand.lower()]
    # partial match
    for cand in candidates:
        for k, v in col_map.items():
            if cand.lower() in k:
                return v
    return None

def read_fast(file_obj, sheet_name=0):
    """Read Excel with dtype=str (fast), then convert week cols to numeric."""
    df = pd.read_excel(file_obj, sheet_name=sheet_name, engine='openpyxl', dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    week_cols = get_week_cols(df)
    for c in week_cols:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    return df


def run_promo_uplift(ly_file, ty_file, promo_start, promo_end, status_cb=None):
    """
    status_cb: optional callback(msg) for progress updates
    Returns: (buf: BytesIO, stats: dict)
    """
    def status(msg):
        if status_cb:
            status_cb(msg)

    promo_start = promo_start.date() if hasattr(promo_start, 'date') else promo_start
    promo_end   = promo_end.date()   if hasattr(promo_end, 'date')   else promo_end

    # ── Load files ─────────────────────────────────────────────
    status("Reading last year actuals...")
    ly_xl = pd.ExcelFile(ly_file, engine='openpyxl')
    ly_actual_sheet = next((s for s in ly_xl.sheet_names if 'actual' in s.lower()), ly_xl.sheet_names[0])
    ly = read_fast(ly_file, sheet_name=ly_actual_sheet)

    status("Reading this year forecast...")
    ty_xl = pd.ExcelFile(ty_file, engine='openpyxl')
    ty_fc_sheet = next((s for s in ty_xl.sheet_names if 'forecast' in s.lower()), None)
    if not ty_fc_sheet:
        raise ValueError("No forecast sheet found in this year's file.")
    ty_fc = read_fast(ty_file, sheet_name=ty_fc_sheet)

    status("Detecting columns...")

    # ── Detect week cols ────────────────────────────────────────
    ly_week_cols = get_week_cols(ly)
    ty_week_cols = get_week_cols(ty_fc)

    # ── Key dimension columns ───────────────────────────────────
    ly_subtype_col = find_col(ly,    ['subtype', 'sub type', 'type'])
    ly_country_col = find_col(ly,    ['country', 'market'])
    ty_subtype_col = find_col(ty_fc, ['subtype', 'sub type', 'type'])
    ty_country_col = find_col(ty_fc, ['country', 'market'])

    if not all([ly_subtype_col, ly_country_col, ty_subtype_col, ty_country_col]):
        raise ValueError("Could not find Subtype or Country columns in one of the files.")

    # Optional output dim cols (only include if they exist)
    OUTPUT_DIMS = ['Chain', 'Pet', 'Subtype', 'ProductDescription', 'ProductID', 'Country']
    ty_output_dims = [(c, find_col(ty_fc, [c])) for c in OUTPUT_DIMS]
    ty_output_dims = [(label, col) for label, col in ty_output_dims if col is not None]

    # ── Split weeks into promo vs normal ────────────────────────
    ty_promo_cols  = [c for c in ty_week_cols if parse_date(c) and promo_start <= parse_date(c) <= promo_end]
    ty_normal_cols = [c for c in ty_week_cols if parse_date(c) and not (promo_start <= parse_date(c) <= promo_end)]

    if not ty_promo_cols:
        avail = [str(c).strip() for c in ty_week_cols[:5]]
        raise ValueError(f"No forecast weeks found between {promo_start} and {promo_end}. Available: {avail}...")

    # LY equivalent promo period (52 weeks back)
    ly_promo_start = promo_start - timedelta(weeks=52)
    ly_promo_end   = promo_end   - timedelta(weeks=52)
    ly_promo_cols  = [c for c in ly_week_cols if parse_date(c) and ly_promo_start <= parse_date(c) <= ly_promo_end]

    # Fallback: try same dates
    if not ly_promo_cols:
        ly_promo_cols = [c for c in ly_week_cols if parse_date(c) and promo_start <= parse_date(c) <= promo_end]

    if not ly_promo_cols:
        avail = [str(c).strip() for c in ly_week_cols[:3]]
        raise ValueError(
            f"No last year weeks found for {ly_promo_start}–{ly_promo_end}. "
            f"LY file starts at: {avail}"
        )

    status(f"Found {len(ty_promo_cols)} promo weeks, {len(ly_promo_cols)} LY reference weeks...")

    # ── LY: aggregate per Subtype × Country ────────────────────
    status("Calculating last year uplift factors...")
    g_ly = [ly_subtype_col, ly_country_col]

    ly['_avg'] = ly[ly_week_cols].mean(axis=1)
    ly['_promo'] = ly[ly_promo_cols].sum(axis=1) / len(ly_promo_cols)

    ly_agg = ly.groupby(g_ly, as_index=False).agg(
        ly_avg_weekly=('_avg', 'sum'),
        ly_promo_weekly=('_promo', 'sum')
    )
    ly_agg['ly_uplift'] = np.where(
        ly_agg['ly_avg_weekly'] > 0,
        ly_agg['ly_promo_weekly'] / ly_agg['ly_avg_weekly'],
        np.nan
    )

    # ── TY: aggregate per Subtype × Country ────────────────────
    status("Calculating this year tool uplift...")
    g_ty = [ty_subtype_col, ty_country_col]

    ty_fc['_ty_normal'] = ty_fc[ty_normal_cols].mean(axis=1) if ty_normal_cols else 0
    ty_fc['_ty_promo']  = ty_fc[ty_promo_cols].mean(axis=1)

    ty_agg = ty_fc.groupby(g_ty, as_index=False).agg(
        ty_normal=('_ty_normal', 'sum'),
        ty_promo=('_ty_promo', 'sum')
    )
    ty_agg['tool_uplift'] = np.where(
        ty_agg['ty_normal'] > 0,
        ty_agg['ty_promo'] / ty_agg['ty_normal'],
        np.nan
    )

    # ── Merge and calculate net uplift ─────────────────────────
    status("Calculating net uplift per subtype/country...")
    merged = ty_agg.rename(columns={ty_subtype_col: '_sub', ty_country_col: '_cty'}).merge(
        ly_agg.rename(columns={ly_subtype_col: '_sub', ly_country_col: '_cty'}),
        on=['_sub', '_cty'], how='left'
    )
    merged['net_uplift'] = np.where(
        (merged['tool_uplift'].fillna(0) > 0) & merged['ly_uplift'].notna(),
        merged['ly_uplift'] / merged['tool_uplift'],
        merged['ly_uplift']
    )
    merged['additional_uplift'] = (merged['net_uplift'] - 1.0).clip(lower=0)

    uplift_map = {}
    for _, row in merged.iterrows():
        uplift_map[(row['_sub'], row['_cty'])] = {
            'ly':    row['ly_uplift'],
            'tool':  row['tool_uplift'],
            'net':   row['net_uplift'],
            'add':   row['additional_uplift'],
        }

    # ── Build SKU-level output ──────────────────────────────────
    status("Building per-SKU output...")
    ty_fc['_sub'] = ty_fc[ty_subtype_col].astype(str).str.strip()
    ty_fc['_cty'] = ty_fc[ty_country_col].astype(str).str.strip()

    rows = []
    promo_col_names = [str(c).strip() for c in ty_promo_cols]

    for _, row in ty_fc.iterrows():
        key   = (row['_sub'], row['_cty'])
        info  = uplift_map.get(key, {})
        add   = info.get('add', np.nan)

        out = {}
        for label, col in ty_output_dims:
            out[label] = row[col]

        for wc, wname in zip(ty_promo_cols, promo_col_names):
            fc_val = float(row[wc]) if pd.notna(row[wc]) else 0.0
            out[wname] = round(fc_val * add) if pd.notna(add) and add > 0 else 0

        out['LY uplift']   = round(info['ly'],   3) if pd.notna(info.get('ly'))   else ''
        out['Tool uplift'] = round(info['tool'],  3) if pd.notna(info.get('tool')) else ''
        out['Net uplift']  = round(info['net'],   3) if pd.notna(info.get('net'))  else ''
        out['Add. uplift'] = f"{round(add*100,1)}%" if pd.notna(add) else ''
        rows.append(out)

    output_df = pd.DataFrame(rows)

    # ── Summary tab ─────────────────────────────────────────────
    summary = merged[['_sub','_cty','ly_avg_weekly','ly_promo_weekly','ly_uplift',
                       'ty_normal','ty_promo','tool_uplift','net_uplift','additional_uplift']].copy()
    summary.columns = ['Subtype','Country','LY avg weekly','LY promo weekly','LY uplift factor',
                        'TY avg weekly forecast','TY promo weekly forecast','Tool uplift factor',
                        'Net uplift factor','Additional uplift']
    summary['Additional uplift'] = summary['Additional uplift'].apply(
        lambda x: f"{round(x*100,1)}%" if pd.notna(x) else '')
    for col in ['LY uplift factor','Tool uplift factor','Net uplift factor']:
        summary[col] = summary[col].apply(lambda x: round(x,3) if pd.notna(x) else '')

    # ── Write output ────────────────────────────────────────────
    status("Writing Excel output...")
    import gc
    del ly, ty_fc, merged
    gc.collect()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name='Promo Registration', index=False)
        summary.to_excel(writer, sheet_name='Uplift Summary', index=False)

        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets['Promo Registration']
        teal_fill  = PatternFill('solid', fgColor='3BBFBF')
        week_fill  = PatternFill('solid', fgColor='E3F7F7')
        white_bold = Font(bold=True, color='FFFFFF')

        for cell in ws[1]:
            col_name = str(cell.value or '')
            if re.search(r'\d{2}/\d{2}/\d{4}', col_name):
                cell.fill = week_fill
                cell.font = Font(bold=True)
            else:
                cell.fill = teal_fill
                cell.font = white_bold

        for col in ws.columns:
            ml = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 25)

    buf.seek(0)
    stats = {
        'promo_weeks': len(ty_promo_cols),
        'sku_count':   len(output_df),
        'subtypes':    int(merged['_sub'].nunique()) if '_sub' in merged.columns else 0,
        'countries':   int(merged['_cty'].nunique()) if '_cty' in merged.columns else 0,
        'avg_ly_uplift': round(float(merged['ly_uplift'].mean()), 2)
                         if '_sub' in merged.columns and merged['ly_uplift'].notna().any() else None,
    }
    return buf, stats
