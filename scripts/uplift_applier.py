"""
uplift_applier.py
Upload a workbench export (sales + forecast tabs).
Choose an uplift % and promo weeks.
Output: per SKU x Chain x Country — units to register as promo
        = (avg_weekly_sales × uplift_factor) − current_forecast
"""

import pandas as pd
import numpy as np
import io
import re
import gc
from datetime import datetime

KNOWN_DIM_COLS = {
    'chain', 'pet', 'subtype', 'productdescription', 'product description',
    'productid', 'product id', 'country', 'ean', 'material', 'sku', 'description'
}

def find_col(df, candidates):
    col_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in col_map:
            return col_map[cand.lower()]
    for cand in candidates:
        for k, v in col_map.items():
            if cand.lower() in k:
                return v
    return None

def parse_date(col):
    try:
        return datetime.strptime(str(col).strip(), "%d/%m/%Y").date()
    except Exception:
        return None

def get_week_cols(df):
    return [c for c in df.columns if parse_date(c) is not None]

def run_uplift_applier(file_obj, uplift_pct, promo_start, promo_end, status_cb=None):
    """
    Parameters
    ----------
    file_obj    : file-like — workbench export with sales + forecast tabs
    uplift_pct  : float — uplift percentage e.g. 50.0 for +50%
    promo_start : datetime — start of promo period
    promo_end   : datetime — end of promo period

    Returns
    -------
    buf   : BytesIO — Excel output
    stats : dict
    """
    def status(msg):
        if status_cb:
            status_cb(msg)

    promo_start = promo_start.date() if hasattr(promo_start, 'date') else promo_start
    promo_end   = promo_end.date()   if hasattr(promo_end, 'date')   else promo_end
    uplift_factor = 1 + (uplift_pct / 100)

    # ── Detect sheets ──────────────────────────────────────────
    status("Opening file...")
    xl = pd.ExcelFile(file_obj, engine='openpyxl')
    sales_sheet = next((s for s in xl.sheet_names if 'actual' in s.lower() or 'sales' in s.lower()), xl.sheet_names[0])
    fc_sheet    = next((s for s in xl.sheet_names if 'forecast' in s.lower()), None)
    if not fc_sheet:
        raise ValueError(f"No forecast sheet found. Available sheets: {xl.sheet_names}")

    # ── Read sales ─────────────────────────────────────────────
    status("Reading sales data...")
    if hasattr(file_obj, 'seek'): file_obj.seek(0)
    sales = pd.read_excel(file_obj, sheet_name=sales_sheet, engine='openpyxl', dtype=str)
    sales.columns = [str(c).strip() for c in sales.columns]
    sales_week_cols = get_week_cols(sales)
    for c in sales_week_cols:
        sales[c] = pd.to_numeric(sales[c], errors='coerce').fillna(0)

    # ── Read forecast ──────────────────────────────────────────
    status("Reading forecast data...")
    if hasattr(file_obj, 'seek'): file_obj.seek(0)
    fc = pd.read_excel(file_obj, sheet_name=fc_sheet, engine='openpyxl', dtype=str)
    fc.columns = [str(c).strip() for c in fc.columns]
    fc_week_cols = get_week_cols(fc)
    for c in fc_week_cols:
        fc[c] = pd.to_numeric(fc[c], errors='coerce').fillna(0)

    # ── Detect promo weeks in forecast ─────────────────────────
    promo_fc_cols = [c for c in fc_week_cols
                     if parse_date(c) and promo_start <= parse_date(c) <= promo_end]
    if not promo_fc_cols:
        avail = [str(c).strip() for c in fc_week_cols[:5]]
        raise ValueError(
            f"No forecast weeks found between {promo_start} and {promo_end}. "
            f"Available: {avail}..."
        )

    status(f"Found {len(promo_fc_cols)} promo weeks in forecast...")

    # ── Detect key columns ─────────────────────────────────────
    s_chain = find_col(sales, ['chain', 'customer', 'client'])
    s_sku   = find_col(sales, ['productid', 'product id', 'sku', 'material', 'ean'])
    s_cty   = find_col(sales, ['country', 'market'])

    f_chain = find_col(fc, ['chain', 'customer', 'client'])
    f_sku   = find_col(fc, ['productid', 'product id', 'sku', 'material', 'ean'])
    f_cty   = find_col(fc, ['country', 'market'])

    for col, name in [(s_chain,'Chain in sales'), (s_sku,'ProductID in sales'),
                      (s_cty,'Country in sales'), (f_chain,'Chain in forecast'),
                      (f_sku,'ProductID in forecast'), (f_cty,'Country in forecast')]:
        if not col:
            raise ValueError(f"{name} column not found.")

    # Output dim cols (only include if they exist in forecast)
    OUTPUT_DIMS = ['Chain', 'Pet', 'Subtype', 'ProductDescription', 'ProductID', 'Country']
    out_dims = [(c, find_col(fc, [c])) for c in OUTPUT_DIMS]
    out_dims = [(label, col) for label, col in out_dims if col]

    # ── Calculate avg weekly sales per SKU x Chain x Country ───
    status("Calculating average weekly sales...")
    n_weeks = len(sales_week_cols)
    if n_weeks == 0:
        raise ValueError("No week columns found in sales tab.")

    sales['_avg'] = sales[sales_week_cols].mean(axis=1)
    g_sales = [s_chain, s_sku, s_cty]
    sales_agg = sales.groupby(g_sales, as_index=False).agg(avg_weekly=('_avg', 'sum'))
    sales_agg['target_per_week'] = sales_agg['avg_weekly'] * uplift_factor
    sales_agg = sales_agg.rename(columns={s_chain: '_chain', s_sku: '_sku', s_cty: '_cty'})

    del sales; gc.collect()

    # ── Build output ───────────────────────────────────────────
    status("Calculating promo units to register...")
    fc['_chain'] = fc[f_chain].astype(str).str.strip()
    fc['_sku']   = fc[f_sku].astype(str).str.strip()
    fc['_cty']   = fc[f_cty].astype(str).str.strip()

    sales_map = {
        (row['_chain'], row['_sku'], row['_cty']): row['target_per_week']
        for _, row in sales_agg.iterrows()
    }

    promo_col_names = [str(c).strip() for c in promo_fc_cols]
    rows = []

    for _, row in fc.iterrows():
        key    = (row['_chain'], row['_sku'], row['_cty'])
        target = sales_map.get(key, np.nan)

        out = {label: row[col] for label, col in out_dims}
        out['Avg weekly sales'] = round(target / uplift_factor, 1) if pd.notna(target) else ''
        out[f'Target ({uplift_pct:g}% uplift)'] = round(target, 1) if pd.notna(target) else ''

        total_to_register = 0
        for wc, wname in zip(promo_fc_cols, promo_col_names):
            fc_val  = float(row[wc]) if pd.notna(row.get(wc)) else 0.0
            if pd.notna(target):
                to_register = max(0, round(target - fc_val))
            else:
                to_register = 0
            out[f'FC {wname}']       = round(fc_val)
            out[f'Register {wname}'] = to_register
            total_to_register += to_register

        out['Total to register'] = total_to_register
        rows.append(out)

    output_df = pd.DataFrame(rows)

    # Filter out rows where nothing needs to be registered
    has_registration = output_df['Total to register'] > 0
    output_filtered  = output_df[has_registration].copy()
    output_all       = output_df.copy()

    del fc, sales_agg; gc.collect()

    # ── Write Excel ────────────────────────────────────────────
    status("Writing Excel output...")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Sheet 1: only rows with something to register
        output_filtered.to_excel(writer, sheet_name='To Register', index=False)
        # Sheet 2: all rows
        output_all.to_excel(writer, sheet_name='All SKUs', index=False)

        from openpyxl.styles import PatternFill, Font
        for sheet_name in ['To Register', 'All SKUs']:
            ws = writer.sheets[sheet_name]
            teal_fill  = PatternFill('solid', fgColor='3BBFBF')
            green_fill = PatternFill('solid', fgColor='E3F7F7')
            amber_fill = PatternFill('solid', fgColor='FEF3E2')
            white_bold = Font(bold=True, color='FFFFFF')

            for cell in ws[1]:
                col_name = str(cell.value or '')
                if col_name.startswith('Register'):
                    cell.fill = amber_fill
                    cell.font = Font(bold=True, color='7A5010')
                elif col_name.startswith('FC'):
                    cell.fill = green_fill
                    cell.font = Font(bold=True)
                else:
                    cell.fill = teal_fill
                    cell.font = white_bold

            for col in ws.columns:
                ml = max((len(str(c.value or '')) for c in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(ml + 3, 25)

    buf.seek(0)

    stats = {
        'promo_weeks':     len(promo_fc_cols),
        'total_skus':      len(output_all),
        'skus_to_register': int(has_registration.sum()),
        'uplift_pct':      uplift_pct,
        'avg_sales_weeks': n_weeks,
    }
    return buf, stats
